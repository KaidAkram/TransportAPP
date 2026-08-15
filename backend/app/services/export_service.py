import io
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def export_to_excel(title: str, headers: list[str], rows: list[list[any]]) ->bytes:
  wb = openpyxl.Workbook()
  ws = wb.active
  ws.title = title[:31] # Excel tab name max 31 chars

  # 1. Company Header
  ws.merge_cells("A1:G1")
  title_cell = ws["A1"]
  title_cell.value = f"E-TRANSPORT ERP — {title.upper()}"
  title_cell.font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
  title_cell.fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
  title_cell.alignment = Alignment(horizontal="center", vertical="center")
  ws.row_dimensions[1].height = 30

  # 2. Date Meta
  ws.merge_cells("A2:G2")
  date_cell = ws["A2"]
  date_cell.value = f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')} | Document Officiel d'Exploitation"
  date_cell.font = Font(name="Calibri", size=9, italic=True, color="6B7280")
  date_cell.alignment = Alignment(horizontal="center", vertical="center")
  ws.row_dimensions[2].height = 18

  # 3. Table Column Headers
  header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
  header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
  thin_border = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
  )

  ws.append([]) # Row 3 blank
  ws.append(headers) # Row 4
  header_row_idx = 4
  ws.row_dimensions[header_row_idx].height = 24

  for col_idx in range(1, len(headers) + 1):
    cell = ws.cell(row=header_row_idx, column=col_idx)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border

  # 4. Data Rows
  row_alt_fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")

  for r_idx, row_data in enumerate(rows, start=5):
    ws.append(row_data)
    ws.row_dimensions[r_idx].height = 20
    is_alt = (r_idx % 2 == 0)

    for col_idx in range(1, len(row_data) + 1):
      cell = ws.cell(row=r_idx, column=col_idx)
      cell.font = Font(name="Calibri", size=10)
      cell.border = thin_border
      if is_alt:
        cell.fill = row_alt_fill

      # Align numbers to right, strings to left
      if isinstance(cell.value, (int, float)):
        cell.alignment = Alignment(horizontal="right", vertical="center")
        if isinstance(cell.value, float):
          cell.number_format = "#,##0.00"
      else:
        cell.alignment = Alignment(horizontal="left", vertical="center")

  # 5. Auto-fit column widths
  for col in ws.columns:
    max_len = 0
    col_letter = get_column_letter(col[0].column)
    for cell in col:
      if cell.row >2 and cell.value:
        max_len = max(max_len, len(str(cell.value)))
    ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

  output = io.BytesIO()
  wb.save(output)
  output.seek(0)
  return output.getvalue()
